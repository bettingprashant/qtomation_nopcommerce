import openpyxl
from openpyxl.styles import PatternFill

def get_row_count(file, sheetname):
    """
    Returns the number of rows in a given sheet of an Excel file.

    Args:
        file (str): The path to the Excel file.
        sheetname (str): The name of the sheet.

    Returns:
        int: The number of rows in the sheet.
    """
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_row

def get_column_count(file, sheetname):
    """
    Returns the number of columns in a given sheet of an Excel file.

    Args:
        file (str): The path to the Excel file.
        sheetname (str): The name of the sheet.

    Returns:
        int: The number of columns in the sheet.
    """
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_column

def read_data(file, sheetname, row_num, column_num):
    """
    Reads data from a specific cell in a given sheet of an Excel file.

    Args:
        file (str): The path to the Excel file.
        sheetname (str): The name of the sheet.
        row_num (int): The row number of the cell (1-based).
        column_num (int): The column number of the cell (1-based).

    Returns:
        any: The value of the cell.
    """
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(row=row_num, column=column_num).value

def write_data(file, sheetname, row_num, column_num, data):
    """
    Writes data to a specific cell in a given sheet of an Excel file.

    Args:
        file (str): The path to the Excel file.
        sheetname (str): The name of the sheet.
        row_num (int): The row number of the cell (1-based).
        column_num (int): The column number of the cell (1-based).
        data (any): The data to write to the cell.
    """
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row=row_num, column=column_num).value = data
    workbook.save(file)